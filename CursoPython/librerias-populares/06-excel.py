# pipenv install openpyxl

import openpyxl

wb = openpyxl.load_workbook("/media/WD/PROYECTOS/2023/python_big_data_and_ai/python_big_data_and_ai/CursoPython/librerias-populares/planilla.xlsx")
print(wb.sheetnames)
hoja1 = wb["Hoja1"]
print(hoja1)
hoja = wb.active
print(hoja)

wb.create_sheet("Hoja3")
print(wb.sheetnames)

hoja3 = wb["Hoja3"]
hoja3.title = "Nuevo titulo"

print(
    hoja1.max_row,
    hoja1.max_column
)

celda = hoja1["A1"]
print(celda.value)
celda.value = "nombre completo"
print(celda.value)

#otra forma
celda2 = hoja1.cell(row=2, column=1) #no parten de 0, el indice parte de 1,2,3....
print(
    celda2.value,
    celda2.row,
    celda2.column,
    celda2.coordinate
    )

for fila in range(1, hoja1.max_row + 1):
    for columna in range(1, hoja1.max_column):
        celda = hoja1.cell(row=fila, column=columna)
        print(fila, columna,celda.value)

#obtener una columna
columna = hoja1["A"]
print(columna)
fila = hoja1["1"]
print(fila)

hoja1.append([1,2,3])
print(hoja1.rows)

hoja1.delete_rows(1,1)

wb.save("/media/WD/PROYECTOS/2023/python_big_data_and_ai/python_big_data_and_ai/CursoPython/librerias-populares/planillaNueva.xlsx")